#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Description:  Converts course data from xlsx to json format.
#               Input: *.xlsx (arg 1)
#				Output: *.json (arg 2)
#				Example usage: py -3 kursplaner_to_json.py kursplaner_miun.xlsx
# Created by:   Jimmy Åhlander
# Date:         2024-03-19
# Modified:     2024-05-07

from openpyxl import load_workbook, Workbook
import json
import sys

# Constants
UNIVERSITY = "Mittuniversitetet"

def date_to_semester(date):
	if date[5:7] < "07": return date[0:4] + ":1" # YYYY:1
	return date[0:4] + ":2" # YYYY:2

def ILOs_to_list(ILOs_raw, ignore_missing_ILOs = False):
	# ILOs may be separated by "-", "•", "  " or use an enumerated list, i.e. "1.", "2.", and so on.
	ILOs = []
	if ILOs_raw:
		ILOs = ILOs_raw.split("  ")
		ILOs = [x for x in ILOs if x]
	for ILO in ILOs:
		print(ILO)
	return ILOs

def get_course_type_from_row(row):
	# vidareutbildning/uppdragsutbildning/förberedande utbildning/grundutbildning/forskarutbildning
	if "U" in row[0].value[5]: return "Uppdragsutbildning"
	if "X" in row[0].value[5]: return "Förberedande utbildning" # there are also preparatory courses with format AB123H, but H-courses give hp, not fup
	if "GR" in row[1].value: return "Grundutbildning"
	if "AV" in row[1].value: return "Grundutbildning"
	if "FO" in row[1].value: return "Forskarutbildning"
	return ""

def main(input_filename, output_filename):
	wb = load_workbook(filename = input_filename)
	ws = wb.active

	courses = []
	course = {}

	for row in ws.iter_rows(min_row=2):
		# grundläggande- och avancerad nivå, forskarnivå, samt uppdragsutbildning
		# exkluderar nivå X vilket påverkar en enda kurs
		if (
				"AV" in row[1].value or "GR" in row[1].value or
				"FO" in row[1].value or "BE" in row[1].value
		   ):
			course = {}
			course['University'] = UNIVERSITY
			course['CourseCode'] = row[0].value
			course['ECTS-credits'] = float(row[3].value)
			ValidFromDate = row[4].value
			course['ValidFrom'] = date_to_semester(ValidFromDate)
			ILOs_sv_raw = row[10].value
			# ILOs_to_list(ILOs_sv_raw)
			course['ILO-sv'] = ILOs_sv_raw if ILOs_sv_raw else ""
			ILOs_en_raw = row[11].value
			course['ILO-en'] = ILOs_en_raw if ILOs_en_raw else ""
			course['SCB-ID'] = row[7].value[0:3] if row[7].value[0:3] else ""
			course['CourseLevel-ID'] = row[9].value[0:3] if row[9].value[0:3] else ""
			course['Prerequisites-sv'] = row[12].value if row[12].value else ""
			course['Prerequisites-en'] = "" # missing in input data
			course['CourseType'] = get_course_type_from_row(row)

			courses.append(course)
			#print(course)

	with open(output_filename, 'w', encoding='utf8') as json_file:
		json.dump(courses, json_file, ensure_ascii=False)

if __name__ == "__main__":
	input_filename = sys.argv[1] if len(sys.argv) > 1 else None
	output_filename = sys.argv[2] if len(sys.argv) > 2 else "result.json"

	if input_filename:
		main(input_filename, output_filename)
	else:
		print("ERROR: Missing input file", file=sys.stderr)